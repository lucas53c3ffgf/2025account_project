import json
import os
import secrets
import smtplib
import sqlite3
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from pathlib import Path
import io
import csv

from flask import Flask, redirect, render_template, request, session, url_for, Response, flash

app = Flask(__name__)
app.secret_key = "dev-secret-key-change-me"

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DB_FILE = DATA_DIR / "app.db"
LEGACY_USERS_FILE = DATA_DIR / "users.json"

BRANCHES = ["Headquarters", "East Branch", "West Branch"]
STATUS_OPTIONS = ["Active", "Inactive"]
DEFAULT_AVG_HOURS_PER_WEEK = 40.0
BRANCH_LOOKUP = {branch.lower(): branch for branch in BRANCHES}
STATUS_LOOKUP = {status.lower(): status for status in STATUS_OPTIONS}
DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S"
PASSWORD_RESET_TOKEN_TTL_MINUTES = 60


# Database helpers.
def get_db_connection():
    conn = sqlite3.connect(DB_FILE, timeout=30)
    conn.row_factory = sqlite3.Row
    return conn          


def migrate_legacy_users_if_needed(conn):
    user_count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    # Skip this if users are already in the database.
    if user_count != 0 or not LEGACY_USERS_FILE.exists():
        return

    try:
        legacy_data = json.loads(LEGACY_USERS_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return

    if not isinstance(legacy_data, dict):
        return

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for email, password in legacy_data.items():
        if isinstance(email, str) and isinstance(password, str):
            # Save emails in lowercase for consistency.
            conn.execute(
                "INSERT OR IGNORE INTO users (email, password, created_at) VALUES (?, ?, ?)",
                (email.strip().lower(), password, now),
            )


def init_db():
    DATA_DIR.mkdir(exist_ok=True)

    with get_db_connection() as conn: 
        # Create the tables the app needs.
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                email TEXT PRIMARY KEY,
                password TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS password_reset_tokens (
                token TEXT PRIMARY KEY,
                email TEXT NOT NULL,
                created_at TEXT NOT NULL,
                expires_at TEXT NOT NULL,
                used_at TEXT
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS employees (
                owner_email TEXT NOT NULL DEFAULT '',
                id TEXT NOT NULL,
                name TEXT NOT NULL,
                branch TEXT NOT NULL,
                hourly_rate REAL NOT NULL,
                phone_number TEXT NOT NULL,
                status TEXT NOT NULL,
                PRIMARY KEY (owner_email, id)
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS change_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT NOT NULL,
                owner_email TEXT NOT NULL DEFAULT '',
                employee_id TEXT NOT NULL,
                employee_name TEXT NOT NULL,
                branch TEXT NOT NULL,
                change_type TEXT NOT NULL,
                old_rate TEXT NOT NULL,
                new_rate TEXT NOT NULL,
                details TEXT NOT NULL,
                changed_by TEXT NOT NULL
            )
            """
        )

        employee_columns = [row["name"] for row in conn.execute("PRAGMA table_info(employees)").fetchall()]
        if "owner_email" not in employee_columns:
            conn.execute("ALTER TABLE employees ADD COLUMN owner_email TEXT NOT NULL DEFAULT ''")

        # Fix older employee tables that used only `id` as the main key.
        employee_pk_cols = [
            row["name"] for row in conn.execute("PRAGMA table_info(employees)").fetchall() if row["pk"] > 0
        ]
        if employee_pk_cols == ["id"]:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS employees_v2 (
                    owner_email TEXT NOT NULL,
                    id TEXT NOT NULL,
                    name TEXT NOT NULL,
                    branch TEXT NOT NULL,
                    hourly_rate REAL NOT NULL,
                    phone_number TEXT NOT NULL,
                    status TEXT NOT NULL,
                    PRIMARY KEY (owner_email, id)
                )
                """
            )
            conn.execute(
                """
                INSERT OR IGNORE INTO employees_v2 (owner_email, id, name, branch, hourly_rate, phone_number, status)
                SELECT owner_email, id, name, branch, hourly_rate, phone_number, status
                FROM employees
                """
            )
            conn.execute("DROP TABLE employees")
            conn.execute("ALTER TABLE employees_v2 RENAME TO employees")

        history_columns = [row["name"] for row in conn.execute("PRAGMA table_info(change_history)").fetchall()]
        if "owner_email" not in history_columns:
            conn.execute("ALTER TABLE change_history ADD COLUMN owner_email TEXT NOT NULL DEFAULT ''")

        migrate_legacy_users_if_needed(conn)

        # Remove old rows that do not belong to a real user.
        conn.execute(
            "DELETE FROM employees WHERE owner_email = '' OR owner_email NOT IN (SELECT email FROM users)"
        )
        conn.execute(
            "DELETE FROM change_history WHERE owner_email = '' OR owner_email NOT IN (SELECT email FROM users)"
        )
        conn.commit()


# User account helpers.
def find_user(email):
    # Look up one user by email.
    with get_db_connection() as conn:
        row = conn.execute("SELECT email, password FROM users WHERE email = ?", (email,)).fetchone()
    return row


def create_user(email, password):
    # Add a new user to the database.
    now = datetime.now().strftime(DATETIME_FORMAT)
    with get_db_connection() as conn:
        conn.execute(
            "INSERT INTO users (email, password, created_at) VALUES (?, ?, ?)",
            (email, password, now),
        )
        conn.commit()


def update_user_password(email, password):
    # Save the new password for this user.
    with get_db_connection() as conn:
        conn.execute("UPDATE users SET password = ? WHERE email = ?", (password, email))
        conn.commit()


def get_password_reset_token_row(token):
    # Get one reset token from the database.
    with get_db_connection() as conn:
        row = conn.execute(
            """
            SELECT token, email, created_at, expires_at, used_at
            FROM password_reset_tokens
            WHERE token = ?
            """,
            (token,),
        ).fetchone()
    return row


def get_valid_password_reset_token_row(token):
    row = get_password_reset_token_row(token)
    # Return nothing if the token is missing, used, or expired.
    if row is None:
        return None

    if row["used_at"]:
        return None

    try:
        expires_at = datetime.strptime(row["expires_at"], DATETIME_FORMAT)
    except ValueError:
        return None

    if datetime.now() > expires_at:
        return None

    return row


def create_password_reset_token(email):
    # Create a new password reset token.
    token = secrets.token_urlsafe(32)
    created_at = datetime.now()
    expires_at = created_at + timedelta(minutes=PASSWORD_RESET_TOKEN_TTL_MINUTES)

    with get_db_connection() as conn:
        conn.execute(
            """
            INSERT INTO password_reset_tokens (token, email, created_at, expires_at, used_at)
            VALUES (?, ?, ?, ?, NULL)
            """,
            (
                token,
                email,
                created_at.strftime(DATETIME_FORMAT),
                expires_at.strftime(DATETIME_FORMAT),
            ),
        )
        conn.commit()

    return token


def mark_password_reset_token_used(token):
    # Mark the reset token as already used.
    used_at = datetime.now().strftime(DATETIME_FORMAT)
    with get_db_connection() as conn:
        conn.execute("UPDATE password_reset_tokens SET used_at = ? WHERE token = ?", (used_at, token))
        conn.commit()


def get_smtp_config():
    # Read email settings from environment variables.
    host = os.environ.get("SMTP_HOST", "").strip()
    if not host:
        return None

    port_raw = os.environ.get("SMTP_PORT", "587").strip()
    try:
        port = int(port_raw)
    except ValueError:
        port = 587

    username = os.environ.get("SMTP_USERNAME", "").strip()
    password = os.environ.get("SMTP_PASSWORD", "")
    from_email = os.environ.get("SMTP_FROM", "").strip() or username

    use_tls = os.environ.get("SMTP_USE_TLS", "1").strip().lower() not in {"0", "false", "no"}
    use_ssl = os.environ.get("SMTP_USE_SSL", "0").strip().lower() in {"1", "true", "yes"}

    if not from_email:
        return None

    return {
        "host": host,
        "port": port,
        "username": username or None,
        "password": password or None,
        "from_email": from_email,
        "use_tls": use_tls,
        "use_ssl": use_ssl,
    }


def send_email(to_email, subject, body):
    # Send one email message.
    cfg = get_smtp_config()
    if cfg is None:
        return False, "SMTP is not configured"

    msg = MIMEText(body, _charset="utf-8")
    msg["Subject"] = subject
    msg["From"] = cfg["from_email"]
    msg["To"] = to_email

    try:
        smtp_cls = smtplib.SMTP_SSL if cfg["use_ssl"] else smtplib.SMTP
        with smtp_cls(cfg["host"], cfg["port"], timeout=15) as server:
            if cfg["use_tls"] and not cfg["use_ssl"]:
                server.ehlo()
                server.starttls()
                server.ehlo()
            if cfg["username"] and cfg["password"]:
                server.login(cfg["username"], cfg["password"])
            server.send_message(msg)
    except Exception as exc:
        return False, str(exc)

    return True, None


def send_password_reset_email(to_email, reset_link):
    # Build and send the password reset email.
    subject = "Reset your password"
    body = (
        "We received a request to reset your password.\n\n"
        f"Reset link: {reset_link}\n\n"
        f"This link expires in {PASSWORD_RESET_TOKEN_TTL_MINUTES} minutes.\n"
        "If you did not request a password reset, you can ignore this email.\n"
    )
    return send_email(to_email, subject, body)


# Employee and history helpers.
def build_branch_counts(owner_email):
    # Count how many employees are in each branch.
    counts = {branch: 0 for branch in BRANCHES}
    with get_db_connection() as conn:
        rows = conn.execute(
            "SELECT branch, COUNT(*) AS branch_count FROM employees WHERE owner_email = ? GROUP BY branch",
            (owner_email,),
        ).fetchall()

    for row in rows:
        counts[row["branch"]] = row["branch_count"]
    return counts


def total_employee_count(owner_email):
    # Count all employees for this user.
    with get_db_connection() as conn:
        count = conn.execute("SELECT COUNT(*) FROM employees WHERE owner_email = ?", (owner_email,)).fetchone()[0]
    return count


def get_employees(owner_email, selected_branch, search_query):
    # Get employees for one user, with optional filters.
    sql = "SELECT id, name, branch, hourly_rate, phone_number, status FROM employees WHERE owner_email = ?"
    params = [owner_email]

    if selected_branch != "All":
        sql += " AND branch = ?"
        params.append(selected_branch)

    if search_query:
        sql += " AND lower(name) LIKE ?"
        params.append(f"%{search_query}%")

    sql += " ORDER BY id"

    with get_db_connection() as conn:
        rows = conn.execute(sql, params).fetchall()

    return [dict(row) for row in rows]

 
def find_employee(employee_id, owner_email):
    # Get one employee by ID.
    with get_db_connection() as conn:
        row = conn.execute(
            "SELECT id, name, branch, hourly_rate, phone_number, status FROM employees WHERE id = ? AND owner_email = ?",
            (employee_id, owner_email),
        ).fetchone()
    return dict(row) if row else None


def generate_employee_id(owner_email):
    with get_db_connection() as conn:
        rows = conn.execute("SELECT id FROM employees WHERE owner_email = ?", (owner_email,)).fetchall()

    numeric_ids = []
    for row in rows:
        emp_id = row["id"]
        if isinstance(emp_id, str) and emp_id.startswith("E") and emp_id[1:].isdigit():
            numeric_ids.append(int(emp_id[1:]))

    # Make the next employee ID like E001, E002, and so on.
    return f"E{(max(numeric_ids, default=0) + 1):03d}"


def log_change(employee, change_type, old_rate, new_rate, details="", conn=None):
    owns_connection = conn is None
    # Save a change in the history table.
    db_conn = conn if conn is not None else get_db_connection()

    db_conn.execute(
        """
        INSERT INTO change_history (
            timestamp, owner_email, employee_id, employee_name, branch,
            change_type, old_rate, new_rate, details, changed_by
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            session.get("user_email", ""),
            employee.get("id", "-"),
            employee.get("name", "-"),
            employee.get("branch", "-"),
            change_type,
            old_rate,
            new_rate,
            details,
            session.get("user_name", "User"),
        ),
    )

    if owns_connection:
        db_conn.commit()
        db_conn.close()


def get_history_entries(owner_email, search_query, change_type_filter):
    # Get history rows using the selected filters.
    sql = (
        "SELECT timestamp, employee_id, employee_name, branch, "
        "change_type, old_rate, new_rate, details, changed_by "
        "FROM change_history"
    )
    params = [owner_email]
    conditions = ["owner_email = ?"]

    if search_query:
        like_value = f"%{search_query}%"
        conditions.append(
            "("
            "lower(employee_name) LIKE ?"
            " OR lower(employee_id) LIKE ?"
            " OR lower(branch) LIKE ?"
            " OR lower(change_type) LIKE ?"
            ")"
        )
        params.extend([like_value, like_value, like_value, like_value])

    if change_type_filter:
        conditions.append("change_type = ?")
        params.append(change_type_filter)

    if conditions:
        sql += " WHERE " + " AND ".join(conditions)

    sql += " ORDER BY id DESC"

    with get_db_connection() as conn:
        rows = conn.execute(sql, params).fetchall()

    return [dict(row) for row in rows]


def get_history_summary(owner_email):
    # Count how many changes were added, updated, or removed.
    with get_db_connection() as conn:
        rows = conn.execute(
            "SELECT change_type, COUNT(*) AS cnt FROM change_history WHERE owner_email = ? GROUP BY change_type",
            (owner_email,),
        ).fetchall()
        total = conn.execute("SELECT COUNT(*) FROM change_history WHERE owner_email = ?", (owner_email,)).fetchone()[0]

    summary = {"total": total, "updated": 0, "added": 0, "removed": 0}
    for row in rows:
        change_type = row["change_type"]
        if change_type == "Updated":
            summary["updated"] = row["cnt"]
        elif change_type == "Added":
            summary["added"] = row["cnt"]
        elif change_type == "Removed":
            summary["removed"] = row["cnt"]

    return summary


def get_recent_rate_changes(owner_email, limit=10):
    # Get recent pay changes for the dashboard.
    with get_db_connection() as conn:
        rows = conn.execute(
            """
            SELECT timestamp, employee_name, branch, old_rate, new_rate
            FROM change_history
            WHERE owner_email = ?
            AND change_type IN ('Updated', 'Added')
            ORDER BY id DESC
            LIMIT ?
            """,
            (owner_email, limit),
        ).fetchall()

    recent = []
    for row in rows:
        recent.append(
            {
                "employee": row["employee_name"],
                "branch": row["branch"],
                "new_rate": row["new_rate"],
                "old_rate": row["old_rate"],
                "date": row["timestamp"].split(" ")[0],
            }
        )
    return recent


def find_employee_for_report(employee_input, owner_email):
    # Find an employee by ID or by name for the report page.
    cleaned = employee_input.strip()
    if not cleaned:
        return None

    with get_db_connection() as conn:
        row = conn.execute(
            """
            SELECT id, name, branch, hourly_rate, phone_number, status
            FROM employees
            WHERE owner_email = ?
            AND (lower(id) = lower(?) OR lower(name) = lower(?))
            ORDER BY id
            LIMIT 1
            """,
            (owner_email, cleaned, cleaned),
        ).fetchone()

    return dict(row) if row else None


def build_report_metrics(employee, year_int):
    now = datetime.now()
    current_year = now.year
    hourly_rate = float(employee["hourly_rate"])
    avg_hours_per_week = DEFAULT_AVG_HOURS_PER_WEEK
    full_year_hours = avg_hours_per_week * 52

    # Estimate worked hours and pay for the selected year.
    if year_int < current_year:
        total_hours_as_of_now = full_year_hours
    elif year_int == current_year:
        weeks_elapsed = max(1, min(52, (now.timetuple().tm_yday // 7) + 1))
        total_hours_as_of_now = avg_hours_per_week * weeks_elapsed
    else:
        total_hours_as_of_now = 0.0

    projected_end_amount = hourly_rate * full_year_hours

    return {
        "hourly_rate": hourly_rate,
        "average_hours_per_week": avg_hours_per_week,
        "total_hours_as_of_now": round(total_hours_as_of_now, 2),
        "projected_end_amount": round(projected_end_amount, 2),
    }


def get_employee_history_for_year(owner_email, employee_id, year_value):
    # Get change history for one employee in one year.
    with get_db_connection() as conn:
        rows = conn.execute(
            """
            SELECT timestamp, old_rate, new_rate, change_type, details, changed_by
            FROM change_history
            WHERE owner_email = ?
            AND employee_id = ?
            AND timestamp LIKE ?
            ORDER BY id DESC
            """,
            (owner_email, employee_id, f"{year_value}%"),
        ).fetchall()

    return [dict(row) for row in rows]


# Small helper functions.
def require_login_redirect():
    # Send logged-out users back to the login page.
    if not session.get("user_email"):
        return redirect(url_for("login"))
    return None


def safe_next_url(raw_next):
    # Only allow return links that stay inside the employees page.
    if not raw_next:
        return url_for("employees")
    if raw_next.startswith("/employees"):
        return raw_next
    return url_for("employees")


def normalize_excel_value(value):
    # Turn empty Excel values into blank text.
    if value is None:
        return ""
    return str(value).strip()


# Excel import helpers.
def parse_employees_xlsx(file_storage):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None, "Excel import requires `openpyxl`. Install it with: pip install openpyxl"

    try:
        workbook = load_workbook(filename=file_storage, data_only=True)
    except Exception:
        return None, "Unable to read the Excel file. Please upload a valid .xlsx workbook."

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return None, "The Excel file is empty."

    raw_headers = rows[0]
    # Clean up column names so different header styles still work.
    headers = [
        normalize_excel_value(h).lower().replace(" ", "_").replace("-", "_")
        for h in raw_headers
    ]

    required_columns = {"id", "name", "branch", "hourly_rate", "phone_number", "status"}
    missing = [col for col in required_columns if col not in headers]
    if missing:
        return None, f"Missing required columns: {', '.join(sorted(missing))}"

    data_rows = []
    for idx, row in enumerate(rows[1:], start=2):
        if row is None:
            continue

        record = {}
        empty_row = True
        for col_index, header in enumerate(headers):
            cell_value = row[col_index] if col_index < len(row) else None
            normalized = normalize_excel_value(cell_value)
            if normalized:
                empty_row = False
            record[header] = normalized

        # Ignore empty rows in the spreadsheet.
        if empty_row:
            continue

        record["_row_number"] = idx
        data_rows.append(record)

    if not data_rows:
        return None, "No employee rows found in the Excel file."

    return data_rows, None


# Basic page routes.
@app.route('/')
def home():
    # Send logged-in users to the dashboard.
    if session.get("user_email"):
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route('/about')
def about():
    # Show the about page.
    return render_template("about.html")


# Login and password reset routes.
@app.route('/login', methods=["GET", "POST"])
def login():
    error = None
    message = None

    if request.method == "GET" and request.args.get("reset") == "1":
        message = "Your password has been reset. Please log in."

    if request.method == "POST":
        # Save the email in lowercase before checking the account.
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        user = find_user(email)
        if not email or not password:
            error = "Email and password are required."
        elif user is None or user["password"] != password:
            error = "Invalid email or password."
        else:
            session["user_email"] = email
            session["user_name"] = email.split("@")[0].title()
            return redirect(url_for("dashboard"))

    return render_template("login.html", error=error, message=message)


@app.route('/forgot-password', methods=["GET", "POST"])
def forgot_password():
    error = None
    message = None
    dev_reset_link = None

    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()

        if not email:
            error = "Please enter your email."
        else:
            user = find_user(email)
            if user is not None:
                token = create_password_reset_token(email)
                reset_link = url_for("reset_password", token=token, _external=True)

                sent, send_error = send_password_reset_email(email, reset_link)
                if not sent:
                    print(f"[password-reset] Email send failed: {send_error}")
                    print(f"[password-reset] Reset link for {email}: {reset_link}")
                    if app.debug:
                        dev_reset_link = reset_link

            # Show the same message even if the email is not found.
            message = "If an account exists for that email, we sent a password reset link."

    return render_template(
        "forgot_password.html",
        error=error,
        message=message,
        dev_reset_link=dev_reset_link,
    )


@app.route('/reset-password/<token>', methods=["GET", "POST"])
def reset_password(token):
    error = None
    message = None

    # Check that the reset link is still valid.
    token_row = get_valid_password_reset_token_row(token)
    if token_row is None:
        error = "This password reset link is invalid or has expired."
        return render_template(
            "reset_password.html",
            error=error,
            message=message,
            token_valid=False,
            token=token,
        )

    if request.method == "POST":
        # Check the new password before saving it.
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not password or not confirm_password:
            error = "All fields are required."
        elif len(password) < 8:
            error = "Password must be at least 8 characters."
        elif password != confirm_password:
            error = "Passwords do not match."
        else:
            update_user_password(token_row["email"], password)
            mark_password_reset_token_used(token)
            return redirect(url_for("login", reset=1))

    return render_template(
        "reset_password.html",
        error=error,
        message=message,
        token_valid=True,
        token=token,
    )


@app.route('/signup', methods=["GET", "POST"])
def signup():
    error = None
   
    if request.method == "POST":
        # Read the signup form values.
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not email or not password or not confirm_password:
            error = "All fields are required."
        elif password != confirm_password:
            error = "Passwords do not match."
        elif find_user(email) is not None:
            error = "An account with that email already exists."
        else:
            create_user(email, password)
            session["user_email"] = email
            session["user_name"] = email.split("@")[0].title()
            return redirect(url_for("dashboard"))

    return render_template("signup.html", error=error)


# Dashboard and employee routes.
@app.route('/dashboard')
def dashboard():
    # Show dashboard totals and recent changes.
    auth_redirect = require_login_redirect()
    if auth_redirect:
        return auth_redirect

    owner_email = session.get("user_email", "")
    branch_counts = build_branch_counts(owner_email)
    total_employees = total_employee_count(owner_email)

    recent_rate_changes = get_recent_rate_changes(owner_email, limit=10)

    return render_template(
        "dashboard.html",
        user_name=session.get("user_name", "User"),
        total_employees=total_employees,
        branch_counts=branch_counts,
        recent_rate_changes=recent_rate_changes,
    )


@app.route('/employees')
def employees():
    # Show the employee list with filters.
    auth_redirect = require_login_redirect()
    if auth_redirect:
        return auth_redirect

    owner_email = session.get("user_email", "")
    selected_branch = request.args.get("branch", "All")
    search_query = request.args.get("q", "").strip().lower()

    filtered_employees = get_employees(owner_email, selected_branch, search_query)
    branch_counts = build_branch_counts(owner_email)
    total_employees = total_employee_count(owner_email)

    return render_template(
        "employees.html",
        user_name=session.get("user_name", "User"),
        total_employees=total_employees,
        branch_counts=branch_counts,
        selected_branch=selected_branch,
        search_query=request.args.get("q", ""),
        employees=filtered_employees,
        current_url=request.full_path.rstrip("?"),
    )


@app.route('/employees/import-excel', methods=["POST"])
def import_employees_excel():
    auth_redirect = require_login_redirect()
    if auth_redirect:
        return auth_redirect

    owner_email = session.get("user_email", "")

    next_url = url_for("employees")
    uploaded_file = request.files.get("excel_file")

    # Make sure a file was uploaded before reading it.
    if uploaded_file is None or uploaded_file.filename is None or uploaded_file.filename.strip() == "":
        flash("Please choose an Excel file (.xlsx) to import.", "danger")
        return redirect(next_url)

    if not uploaded_file.filename.lower().endswith(".xlsx"):
        flash("Invalid file type. Please upload a .xlsx file.", "danger")
        return redirect(next_url)

    records, parse_error = parse_employees_xlsx(uploaded_file)
    if parse_error:
        flash(parse_error, "danger")
        return redirect(next_url)

    inserted_count = 0
    updated_count = 0
    skipped_count = 0
    # Save a few skipped rows to show in the final message.
    skipped_examples = []

    with get_db_connection() as conn:
        for record in records:
            # Check each row before adding or updating it.
            row_num = record["_row_number"]
            employee_id = record.get("id", "").upper()
            name = record.get("name", "")
            branch = record.get("branch", "")
            hourly_rate_raw = record.get("hourly_rate", "")
            phone_number = record.get("phone_number", "")
            status = record.get("status", "")

            if not employee_id or not name or not branch or not hourly_rate_raw or not phone_number or not status:
                skipped_count += 1
                if len(skipped_examples) < 5:
                    skipped_examples.append(f"row {row_num}: missing required values")
                continue

            normalized_branch = BRANCH_LOOKUP.get(branch.lower())
            if normalized_branch is None:
                skipped_count += 1
                if len(skipped_examples) < 5:
                    skipped_examples.append(f"row {row_num}: invalid branch `{branch}`")
                continue

            normalized_status = STATUS_LOOKUP.get(status.lower())
            if normalized_status is None:
                skipped_count += 1
                if len(skipped_examples) < 5:
                    skipped_examples.append(f"row {row_num}: invalid status `{status}`")
                continue

            try:
                hourly_rate = float(hourly_rate_raw)
            except ValueError:
                hourly_rate = -1.0

            if hourly_rate < 0:
                skipped_count += 1
                if len(skipped_examples) < 5:
                    skipped_examples.append(f"row {row_num}: invalid hourly_rate `{hourly_rate_raw}`")
                continue

            existing = conn.execute(
                "SELECT id, owner_email, name, branch, hourly_rate, phone_number, status FROM employees WHERE id = ? AND owner_email = ?",
                (employee_id, owner_email),
            ).fetchone()

            employee_payload = {
                "id": employee_id,
                "name": name,
                "branch": normalized_branch,
                "hourly_rate": round(hourly_rate, 2),
                "phone_number": phone_number,
                "status": normalized_status,
            }

            # Update the employee if the ID already exists.
            if existing is None:
                conn.execute(
                    """
                    INSERT INTO employees (id, owner_email, name, branch, hourly_rate, phone_number, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        employee_payload["id"],
                        owner_email,
                        employee_payload["name"],
                        employee_payload["branch"],
                        employee_payload["hourly_rate"],
                        employee_payload["phone_number"],
                        employee_payload["status"],
                    ),
                )
                inserted_count += 1
                log_change(
                    employee_payload,
                    "Added",
                    "-",
                    f"${employee_payload['hourly_rate']:.2f}",
                    "Imported from Excel",
                    conn=conn,
                )
            else:
                old_rate = float(existing["hourly_rate"])
                old_snapshot = dict(existing)

                conn.execute(
                    """
                    UPDATE employees
                    SET name = ?, branch = ?, hourly_rate = ?, phone_number = ?, status = ?
                    WHERE id = ? AND owner_email = ?
                    """,
                    (
                        employee_payload["name"],
                        employee_payload["branch"],
                        employee_payload["hourly_rate"],
                        employee_payload["phone_number"],
                        employee_payload["status"],
                        employee_payload["id"],
                        owner_email,
                    ),
                )
                updated_count += 1

                changed_fields = []
                if old_snapshot["name"] != employee_payload["name"]:
                    changed_fields.append("name")
                if old_snapshot["branch"] != employee_payload["branch"]:
                    changed_fields.append("branch")
                if float(old_snapshot["hourly_rate"]) != employee_payload["hourly_rate"]:
                    changed_fields.append("hourly_rate")
                if old_snapshot["phone_number"] != employee_payload["phone_number"]:
                    changed_fields.append("phone")
                if old_snapshot["status"] != employee_payload["status"]:
                    changed_fields.append("status")

                details = "Imported from Excel. Updated fields: " + (", ".join(changed_fields) if changed_fields else "none")
                log_change(
                    employee_payload,
                    "Updated",
                    f"${old_rate:.2f}",
                    f"${employee_payload['hourly_rate']:.2f}",
                    details,
                    conn=conn,
                )

        conn.commit()

    summary = (
        "Excel import complete. "
        f"Added: {inserted_count}, Updated: {updated_count}, Skipped: {skipped_count}. "
        "Rows with existing ID are updated, not duplicated."
    )
    if skipped_examples:
        summary += " Examples: " + "; ".join(skipped_examples)

    flash(summary, "success")
    return redirect(next_url)


@app.route('/employees/add', methods=["GET", "POST"])
def add_employee():
    auth_redirect = require_login_redirect()
    if auth_redirect:
        return auth_redirect

    owner_email = session.get("user_email", "")
    next_url = safe_next_url(request.args.get("next") if request.method == "GET" else request.form.get("next"))
    error = None

    if request.method == "POST":
        # Check the form before creating the employee.
        name = request.form.get("name", "").strip()
        branch = request.form.get("branch", "").strip()
        hourly_rate_raw = request.form.get("hourly_rate", "").strip()
        phone_number = request.form.get("phone_number", "").strip()
        status = request.form.get("status", "Active").strip()

        if not name or not branch or not hourly_rate_raw or not phone_number or not status:
            error = "All fields are required."
        elif branch not in BRANCHES:
            error = "Please select a valid branch."
        elif status not in STATUS_OPTIONS:
            error = "Please select a valid status."
        else:
            try:
                hourly_rate = float(hourly_rate_raw)
            except ValueError:
                hourly_rate = -1.0

            if hourly_rate < 0:
                error = "Hourly rate must be a positive number."
            else:
                new_employee = {
                    "id": generate_employee_id(owner_email),
                    "name": name,
                    "branch": branch,
                    "hourly_rate": round(hourly_rate, 2),
                    "phone_number": phone_number,
                    "status": status,
                }

                with get_db_connection() as conn:
                    conn.execute(
                        """
                        INSERT INTO employees (id, owner_email, name, branch, hourly_rate, phone_number, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            new_employee["id"],
                            owner_email,
                            new_employee["name"],
                            new_employee["branch"],
                            new_employee["hourly_rate"],
                            new_employee["phone_number"],
                            new_employee["status"],
                        ),
                    )
                    conn.commit()

                log_change(new_employee, "Added", "-", f"${new_employee['hourly_rate']:.2f}", "Employee created")
                return redirect(next_url)

    return render_template(
        "add_employee.html",
        user_name=session.get("user_name", "User"),
        branches=BRANCHES,
        statuses=STATUS_OPTIONS,
        next_url=next_url,
        suggested_id=generate_employee_id(owner_email),
        error=error,
    )


@app.route('/employees/<employee_id>/edit', methods=["GET", "POST"])
def edit_employee(employee_id):
    auth_redirect = require_login_redirect()
    if auth_redirect:
        return auth_redirect

    owner_email = session.get("user_email", "")
    employee = find_employee(employee_id, owner_email)
    if employee is None:
        return redirect(url_for("employees"))

    error = None

    if request.method == "POST":
        new_id = request.form.get("id", "").strip().upper()
        name = request.form.get("name", "").strip()
        branch = request.form.get("branch", "").strip()
        hourly_rate_raw = request.form.get("hourly_rate", "").strip()
        phone_number = request.form.get("phone_number", "").strip()
        status = request.form.get("status", "").strip()

        if not new_id or not name or not branch or not hourly_rate_raw or not phone_number or not status:
            error = "All fields are required."
        elif branch not in BRANCHES:
            error = "Please select a valid branch."
        elif status not in STATUS_OPTIONS:
            error = "Please select a valid status."
        else:
            try:
                hourly_rate = float(hourly_rate_raw)
            except ValueError:
                hourly_rate = -1.0

            if hourly_rate < 0:
                error = "Hourly rate must be a positive number."
            elif new_id != employee_id and find_employee(new_id, owner_email) is not None:
                error = "Employee ID already exists. Use a unique ID."
            if error is None:
                old_rate = f"${employee['hourly_rate']:.2f}"
                # Keep the old values so the history log can show what changed.
                old_snapshot = employee.copy()

                with get_db_connection() as conn:
                    conn.execute(
                        """
                        UPDATE employees
                        SET id = ?, name = ?, branch = ?, hourly_rate = ?, phone_number = ?, status = ?
                        WHERE id = ? AND owner_email = ?
                        """,
                        (new_id, name, branch, round(hourly_rate, 2), phone_number, status, employee_id, owner_email),
                    )
                    conn.commit()

                updated_employee = {
                    "id": new_id,
                    "name": name,
                    "branch": branch,
                    "hourly_rate": round(hourly_rate, 2),
                    "phone_number": phone_number,
                    "status": status,
                }

                detail_parts = []
                if old_snapshot["name"] != updated_employee["name"]:
                    detail_parts.append("name")
                if old_snapshot["branch"] != updated_employee["branch"]:
                    detail_parts.append("branch")
                if old_snapshot["phone_number"] != updated_employee["phone_number"]:
                    detail_parts.append("phone")
                if old_snapshot["status"] != updated_employee["status"]:
                    detail_parts.append("status")
                if old_snapshot["id"] != updated_employee["id"]:
                    detail_parts.append("id")

                details = "Updated fields: " + (", ".join(detail_parts) if detail_parts else "none")
                log_change(updated_employee, "Updated", old_rate, f"${updated_employee['hourly_rate']:.2f}", details)
                return redirect(url_for("employees", branch=branch))

    return render_template(
        "edit_employee.html",
        user_name=session.get("user_name", "User"),
        employee=employee,
        branches=BRANCHES,
        statuses=STATUS_OPTIONS,
        error=error,
    )


@app.route('/employees/<employee_id>/remove', methods=["GET", "POST"])
def remove_employee(employee_id):
    # Show the remove page and delete after confirmation.
    auth_redirect = require_login_redirect()
    if auth_redirect:
        return auth_redirect

    owner_email = session.get("user_email", "")
    next_url = safe_next_url(request.args.get("next") if request.method == "GET" else request.form.get("next"))
    employee = find_employee(employee_id, owner_email)

    if employee is None:
        return redirect(next_url)

    if request.method == "POST":
        action = request.form.get("action")
        if action == "confirm":
            log_change(employee, "Removed", f"${employee['hourly_rate']:.2f}", "-", "Employee removed")
            with get_db_connection() as conn:
                conn.execute("DELETE FROM employees WHERE id = ? AND owner_email = ?", (employee_id, owner_email))
                conn.commit()
        return redirect(next_url)

    return render_template(
        "confirm_remove.html",
        user_name=session.get("user_name", "User"),
        employee=employee,
        next_url=next_url,
    )


# History and report routes.
@app.route('/history')
def history():
    # Show saved change history.
    auth_redirect = require_login_redirect()
    if auth_redirect:
        return auth_redirect

    owner_email = session.get("user_email", "")
    search_query = request.args.get("q", "").strip().lower()
    selected_change_type = request.args.get("change_type", "").strip()
    if selected_change_type not in {"Updated", "Added", "Removed"}:
        selected_change_type = ""

    history_entries = get_history_entries(owner_email, search_query, selected_change_type)
    summary = get_history_summary(owner_email)

    return render_template(
        "history.html",
        user_name=session.get("user_name", "User"),
        history_entries=history_entries,
        search_query=request.args.get("q", ""),
        selected_change_type=selected_change_type,
        summary=summary,
    )


@app.route('/report', methods=["GET", "POST"])
def report():
    # Build the report page for one employee and year.
    auth_redirect = require_login_redirect()
    if auth_redirect:
        return auth_redirect

    owner_email = session.get("user_email", "")
    error = None
    report_data = None
    employee_query = ""
    year_value = str(datetime.now().year)

    if request.method == "POST":
        employee_query = request.form.get("employee", "").strip()
        year_value = request.form.get("year", "").strip()

        if not employee_query or not year_value:
            error = "Employee and year are required."
        elif not year_value.isdigit() or len(year_value) != 4:
            error = "Please enter a valid 4-digit year."
        else:
            employee = find_employee_for_report(employee_query, owner_email)
            if employee is None:
                error = "No employee found with that name or ID."
            else:
                # Build the report details for the selected employee and year.
                metrics = build_report_metrics(employee, int(year_value))
                year_history = get_employee_history_for_year(owner_email, employee["id"], year_value)

                report_data = {
                    "employee": employee,
                    "employee_id": employee["id"],
                    "year": year_value,
                    "current_rate": f"${metrics['hourly_rate']:.2f}",
                    "average_hours_per_week": f"{metrics['average_hours_per_week']:.2f}",
                    "total_hours_as_of_now": f"{metrics['total_hours_as_of_now']:.2f}",
                    "projected_end_amount": f"${metrics['projected_end_amount']:,.2f}",
                    "year_history": year_history,
                }

    return render_template(
        "report.html",
        user_name=session.get("user_name", "User"),
        employee_query=employee_query,
        year_value=year_value,
        report_data=report_data,
        error=error,
    )


@app.route('/report/download', methods=["POST"])
def download_report():
    # Download the report as a CSV file.
    auth_redirect = require_login_redirect()
    if auth_redirect:
        return auth_redirect

    owner_email = session.get("user_email", "")
    employee_id = request.form.get("employee_id", "").strip()
    year_value = request.form.get("year", "").strip()

    if not employee_id or not year_value or not year_value.isdigit() or len(year_value) != 4:
        return redirect(url_for("report"))

    employee = find_employee(employee_id, owner_email)
    if employee is None:
        return redirect(url_for("report"))

    # Build the report again using the latest saved data.
    metrics = build_report_metrics(employee, int(year_value))

    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow(
        [
            "Employee ID",
            "Employee Name",
            "Branch",
            "Year",
            "Current Hourly Rate",
            "Average Hours Per Week",
            "Total Hours As Of Now",
            "Projected End Amount",
        ]
    )
    writer.writerow(
        [
            employee["id"],
            employee["name"],
            employee["branch"],
            year_value,
            f"{metrics['hourly_rate']:.2f}",
            f"{metrics['average_hours_per_week']:.2f}",
            f"{metrics['total_hours_as_of_now']:.2f}",
            f"{metrics['projected_end_amount']:.2f}",
        ]
    )

    response = Response(csv_buffer.getvalue(), mimetype="text/csv")
    response.headers["Content-Disposition"] = (
        f"attachment; filename=employee_report_{employee['id']}_{year_value}.csv"
    )
    return response


@app.route('/logout')
def logout():
    # Clear the session and log the user out.
    session.clear()
    return redirect(url_for("login"))


init_db()

if __name__ == "__main__":
    app.run(debug=True)
